"""
Migration — Import CSDL_HangHoa_PA3.xlsx vào SQLite

Cấu trúc thực tế:
  A_DanhMuc:  A=MãHĐ | B=Tên | C=Nhóm | D=ĐVT | E-H=Alias1-4 | I=GhiChú
  C_GiaoDich: A=ID | B=MãHĐ | C=Loại(Bán/Nhập) | D=ĐơnGiá | E=Ngày | F=SL | G=SốBG | H=GhiChú
"""
import os
import json
from datetime import datetime
import openpyxl
from db.database import get_connection

EXCEL_PATH = os.getenv("EXCEL_PATH", "./data/CSDL_HangHoa_PA3.xlsx")


def _parse_date(val) -> datetime | None:
    """Chuyển chuỗi ngày dd/mm/yyyy hoặc datetime thành datetime"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.strptime(str(val).strip(), "%d/%m/%Y")
    except Exception:
        return None


def _load_latest_ban_prices(wb) -> dict:
    """
    Đọc C_GiaoDich, lấy đơn giá Bán mới nhất cho mỗi MãHĐ.
    Trả về dict {ma_hd: gia_ban}
    """
    if "C_GiaoDich" not in wb.sheetnames:
        print("[migrate] Không tìm thấy sheet C_GiaoDich — giá sẽ để trống.")
        return {}

    ws = wb["C_GiaoDich"]
    # {ma_hd: (ngay, don_gia)}
    latest = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 4:
            continue
        ma_hd = row[1]
        loai  = row[2]
        gia   = row[3]
        ngay  = _parse_date(row[4]) if len(row) > 4 else None

        if not ma_hd or not loai:
            continue
        loai_str = str(loai).strip().lower()
        if loai_str != "bán" and loai_str != "ban":
            continue
        try:
            gia_val = float(gia) if gia is not None else None
        except (ValueError, TypeError):
            continue
        if gia_val is None or gia_val <= 0:
            continue

        ma = str(ma_hd).strip()
        existing = latest.get(ma)
        if existing is None:
            latest[ma] = (ngay, gia_val)
        else:
            ex_ngay, _ = existing
            # Giữ giá mới nhất; nếu không có ngày → dùng dòng cuối cùng
            if ngay is None or (ex_ngay is not None and ngay > ex_ngay):
                latest[ma] = (ngay, gia_val)

    prices = {ma: gia for ma, (_, gia) in latest.items()}
    print(f"[migrate] Đọc được {len(prices)} giá bán từ C_GiaoDich.")
    return prices


def run_migration(force: bool = False):
    """
    Đọc A_DanhMuc + C_GiaoDich từ Excel và upsert vào bảng products.
    force=True: xóa toàn bộ và import lại từ đầu.
    """
    if not os.path.exists(EXCEL_PATH):
        print(f"[migrate] Không tìm thấy: {EXCEL_PATH} — bỏ qua.")
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except Exception as e:
        print(f"[migrate] Lỗi mở Excel: {e}")
        return

    if "A_DanhMuc" not in wb.sheetnames:
        print("[migrate] Không tìm thấy sheet A_DanhMuc")
        wb.close()
        return

    # Bước 1: Lấy giá bán mới nhất từ C_GiaoDich
    prices = _load_latest_ban_prices(wb)

    ws = wb["A_DanhMuc"]
    conn = get_connection()

    if force:
        conn.execute("DELETE FROM products")
        conn.commit()
        print("[migrate] Đã xóa dữ liệu cũ.")

    inserted = skipped = 0

    # Cột thực tế: A=MãHĐ | B=Tên | C=Nhóm | D=ĐVT | E=Alias1 | F=Alias2 | G=Alias3 | H=Alias4
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 2:
            continue

        ma_hd = row[0]
        ten   = row[1]
        nhom  = row[2] if len(row) > 2 else None
        dvt   = row[3] if len(row) > 3 else None
        # Alias 1–4
        aliases = []
        for i in range(4, 8):
            if len(row) > i and row[i]:
                aliases.append(str(row[i]).strip())

        if not ma_hd or not ten:
            continue
        ma_hd = str(ma_hd).strip()
        if ma_hd.startswith("???") or ma_hd == "":
            skipped += 1
            continue

        gia_ban = prices.get(ma_hd)  # Giá bán mới nhất, None nếu chưa có

        try:
            conn.execute("""
                INSERT INTO products (ma_hd, ten_hang, don_vi, gia_ban, nhom, aliases)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(ma_hd) DO UPDATE SET
                    ten_hang   = excluded.ten_hang,
                    don_vi     = excluded.don_vi,
                    gia_ban    = excluded.gia_ban,
                    nhom       = excluded.nhom,
                    aliases    = excluded.aliases,
                    updated_at = datetime('now','localtime')
            """, (
                ma_hd,
                str(ten).strip(),
                str(dvt).strip() if dvt else None,
                gia_ban,
                str(nhom).strip() if nhom else None,
                json.dumps(aliases, ensure_ascii=False)
            ))
            inserted += 1
        except Exception as e:
            print(f"[migrate] Lỗi dòng {ma_hd}: {e}")

    conn.commit()
    conn.close()
    wb.close()

    # Reload cache
    from bot.search import load_cache
    load_cache()

    print(f"[migrate] Xong: {inserted} sản phẩm import, bỏ qua {skipped} (??? hoặc rỗng)")
